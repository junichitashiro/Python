from pathlib import Path
from openpyxl import load_workbook


def check_excel_file(file_path: str, sheet_name: str = 'Sheet1') -> bool:
    """
    指定されたExcelファイルとシートの存在を確認し、読み込み可能かを検証する関数

    Args:
        file_path (str): チェックするExcelファイルのパス
        sheet_name (str, optional): 存在を確認するシート名（デフォルトは'Sheet1'）

    Returns:
        bool: ファイルとシートが存在し読み込み可能であれば True、そうでなければ False
    """
    file = Path(file_path)

    # ファイルの存在チェック
    if not file.exists() or file.is_dir():
        print(f'エラー: 指定したファイルが存在しないかディレクトリです -> {file_path}')
        return False

    try:
        # ファイルの読み込みチェック
        wb = load_workbook(file_path, data_only=True)

        # シート名の存在チェック
        if sheet_name not in wb.sheetnames:
            print(f'エラー: 指定したシートが存在しません -> {sheet_name}')
            return False

        print('チェック完了: ファイルとシートは有効です')
        return True

    # 読み込み時のエラーハンドリング
    except FileNotFoundError:
        print(f'エラー: ファイルが見つかりません -> {file_path}')
        return False
    except PermissionError:
        print(f'エラー: ファイルのアクセス権がありません -> {file_path}')
        return False
    except Exception as e:
        print(f'エラー: ファイルを読み込めませんでした -> {e}')
        return False
